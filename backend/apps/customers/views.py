import csv
import io
import re
from datetime import date

import requests as http_requests
from django.db import transaction
from django.db.models import QuerySet
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook, load_workbook
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .filters import CustomerFilter
from .models import Customer
from .serializers import CustomerSerializer
from apps.activity.log import log_activity
from apps.activity.models import ActivityLog
from apps.users.permissions import HasCompanyPermission, IsCompanyMember
from apps.users.tenant import filter_queryset_for_current_company


class CustomerViewSet(viewsets.ModelViewSet):
    """Full CRUD for customers in the user's active company."""
    lookup_field = "uuid"

    serializer_class = CustomerSerializer
    required_permission = 'can_manage_customers'
    read_permission = None  # any company member may list/read customers (needed for orders)
    permission_classes = [IsAuthenticated, IsCompanyMember, HasCompanyPermission]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = CustomerFilter
    search_fields = ["name", "company_name", "nip", "city"]
    ordering_fields = [
        "name",
        "company_name",
        "city",
        "distance_km",
        "payment_terms",
        "credit_limit",
        "created_at",
        "updated_at",
        "is_active",
    ]
    ordering = ["-created_at"]

    def get_queryset(self) -> QuerySet:
        qs = Customer.objects.all().order_by("-created_at")
        return filter_queryset_for_current_company(qs, self.request.user)

    def perform_create(self, serializer):
        instance = serializer.save(
            company=self.request.user.current_company,
            user=self.request.user,
        )
        log_activity(
            user=self.request.user,
            action="customer.create",
            status=ActivityLog.STATUS_SUCCESS,
            object_type="customer",
            object_id=instance.name,
        )

    def perform_update(self, serializer):
        instance = serializer.save(
            company=self.request.user.current_company,
            user=self.request.user,
        )
        log_activity(
            user=self.request.user,
            action="customer.update",
            status=ActivityLog.STATUS_SUCCESS,
            object_type="customer",
            object_id=instance.name,
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        name = instance.name
        instance.delete()
        log_activity(
            user=request.user,
            action="customer.delete",
            status=ActivityLog.STATUS_SUCCESS,
            object_type="customer",
            object_id=name,
        )
        from rest_framework import status
        from rest_framework.response import Response
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(
        detail=False, methods=["get"], url_path="lookup-nip",
        permission_classes=[IsAuthenticated, IsCompanyMember],
    )
    def lookup_nip(self, request):
        """GET /api/customers/lookup-nip/?nip=1234567890
        Searches MF Biała Lista first, then KRS, and returns pre-fill data.
        """
        nip = request.query_params.get("nip", "").strip().replace("-", "").replace(" ", "")
        if not nip or not re.match(r"^\d{10}$", nip):
            return Response(
                {"detail": "Podaj poprawny 10-cyfrowy NIP."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        result = self._fetch_from_mf(nip)
        if result:
            return Response({**result, "source": "mf"})

        result = self._fetch_from_krs(nip)
        if result:
            return Response({**result, "source": "krs"})

        return Response(
            {"detail": "Nie znaleziono firmy o podanym NIP w rejestrach publicznych."},
            status=status.HTTP_404_NOT_FOUND,
        )

    def _fetch_from_mf(self, nip: str) -> dict | None:
        """Query MF Biała Lista (Wykaz Podatników VAT)."""
        today = date.today().isoformat()
        url = f"https://wl-api.mf.gov.pl/api/search/nip/{nip}?date={today}"
        try:
            resp = http_requests.get(url, timeout=5, headers={"User-Agent": "MojeSaldoo/1.0"})
            if resp.status_code != 200:
                return None
            data = resp.json()
            subject = data.get("result", {}).get("subject")
            if not subject:
                return None
            address_raw = subject.get("workingAddress") or subject.get("residenceAddress") or ""
            street, postal_code, city = self._parse_pl_address(address_raw)
            name = subject.get("name", "")
            return {
                "name": name,
                "company_name": name,
                "nip": nip,
                "regon": subject.get("regon", ""),
                "street": street,
                "postal_code": postal_code,
                "city": city,
            }
        except Exception:
            return None

    def _fetch_from_krs(self, nip: str) -> dict | None:
        """Query KRS API (Ministerstwo Sprawiedliwości) as fallback."""
        url = f"https://api-krs.ms.gov.pl/api/krs/podmiotPrawny?nip={nip}&rejestr=P&format=json"
        try:
            resp = http_requests.get(url, timeout=5, headers={"User-Agent": "MojeSaldoo/1.0"})
            if resp.status_code != 200:
                return None
            data = resp.json()
            wyniki = data.get("wyniki", [])
            if not wyniki:
                return None
            first = wyniki[0]
            name = first.get("nazwa") or first.get("nazwaSkrocona") or ""
            address_raw = first.get("adres") or ""
            street, postal_code, city = self._parse_pl_address(address_raw)
            return {
                "name": name,
                "company_name": name,
                "nip": nip,
                "regon": first.get("regon", ""),
                "street": street,
                "postal_code": postal_code,
                "city": city,
            }
        except Exception:
            return None

    @staticmethod
    def _parse_pl_address(address: str) -> tuple[str, str, str]:
        """Parse 'ul. Przykładowa 1, 00-001 Warszawa' → (street, postal_code, city)."""
        m = re.match(r"^(.*),\s*(\d{2}-\d{3})\s+(.+)$", address.strip())
        if m:
            return m.group(1).strip(), m.group(2).strip(), m.group(3).strip()
        return address.strip(), "", ""

    _IMPORT_HEADERS = [
        "Nazwa", "Nazwa firmy", "NIP", "Telefon", "Email",
        "Ulica", "Miasto", "Kod pocztowy", "Termin płatności (dni)",
    ]

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        """GET /api/customers/import-template/ — download a blank XLSX template."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Klienci"
        ws.append(self._IMPORT_HEADERS)
        ws.append(["Jan Kowalski", "Piekarnia Kowalski", "1234567890", "600123456",
                   "jan@piekarnia.pl", "ul. Słoneczna 1", "Warszawa", "00-001", 14])

        for col, width in zip("ABCDEFGHI", [30, 30, 14, 16, 30, 30, 20, 12, 22]):
            ws.column_dimensions[col].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="szablon_klienci.xlsx"'
        return resp

    @action(detail=False, methods=["post"], url_path="import", parser_classes=[MultiPartParser])
    def import_customers(self, request):
        """POST /api/customers/import/ — import customers from CSV or XLSX.

        Dedup by name (case-insensitive): update if exists, create if not.
        """
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"detail": "Brak pliku."}, status=status.HTTP_400_BAD_REQUEST)

        dry_run = request.data.get("dry_run", "true").lower() != "false"
        filename = file_obj.name.lower()

        try:
            rows = self._parse_import_file(file_obj, filename)
        except Exception as exc:
            return Response({"detail": f"Nie udało się odczytać pliku: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

        errors = []
        valid_customers = []

        for i, raw in enumerate(rows, start=2):
            row_errors, customer_data = self._validate_import_row(raw, i)
            if row_errors:
                errors.extend(row_errors)
            else:
                valid_customers.append(customer_data)

        company = request.user.current_company
        name_index = {
            c.name.lower(): c
            for c in Customer.objects.filter(company=company)
        }

        if dry_run or errors:
            to_create = [d for d in valid_customers if d["name"].lower() not in name_index]
            to_update = [d for d in valid_customers if d["name"].lower() in name_index]
            return Response({
                "dry_run": True,
                "to_create": len(to_create),
                "to_update": len(to_update),
                "to_skip": 0,
                "valid_count": len(valid_customers),
                "error_count": len(errors),
                "errors": errors,
            })

        _UPDATE_FIELDS = [
            "name", "company_name", "nip", "phone", "email",
            "street", "city", "postal_code", "payment_terms",
        ]
        created = updated = 0

        with transaction.atomic():
            for data in valid_customers:
                existing = name_index.get(data["name"].lower())
                if existing:
                    for field in _UPDATE_FIELDS:
                        setattr(existing, field, data[field])
                    existing.save(update_fields=_UPDATE_FIELDS)
                    updated += 1
                else:
                    Customer.objects.create(
                        company=company,
                        user=request.user,
                        **data,
                    )
                    created += 1

        log_activity(
            user=request.user,
            action="customer.import",
            status=ActivityLog.STATUS_SUCCESS,
            object_type="customer",
            object_id=f"created={created} updated={updated}",
        )
        return Response({
            "dry_run": False,
            "created": created,
            "updated": updated,
            "skipped": 0,
            "error_count": 0,
            "errors": [],
        }, status=status.HTTP_201_CREATED)

    def _parse_import_file(self, file_obj, filename: str) -> list[dict]:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            wb = load_workbook(file_obj, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
            result = []
            for row in rows_iter:
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                result.append({headers[j]: (str(row[j]).strip() if row[j] is not None else "") for j in range(len(headers))})
            return result
        else:
            raw_bytes = file_obj.read()
            text = raw_bytes.decode("utf-8-sig")
            dialect = "excel" if "," in text.splitlines()[0] and ";" not in text.splitlines()[0] else None
            reader = csv.DictReader(io.StringIO(text), delimiter=";" if dialect is None else ",")
            return [{k.strip(): (v or "").strip() for k, v in row.items() if k is not None} for row in reader]

    def _validate_import_row(self, raw: dict, row_num: int) -> tuple[list[dict], dict | None]:
        norm = {k.strip().lower(): v.strip() for k, v in raw.items() if k}

        def get(*keys):
            for k in keys:
                if k.lower() in norm:
                    return norm[k.lower()]
            return ""

        errors = []

        name = get("nazwa")
        if not name:
            errors.append({"row": row_num, "field": "Nazwa", "message": "Pole wymagane."})

        payment_terms_raw = get("termin płatności (dni)", "termin platnosci (dni)", "termin")
        payment_terms = 14
        if payment_terms_raw:
            try:
                payment_terms = int(payment_terms_raw)
                if payment_terms < 0:
                    raise ValueError
            except ValueError:
                errors.append({"row": row_num, "field": "Termin płatności (dni)", "message": f"Nieprawidłowa wartość: '{payment_terms_raw}'."})

        if errors:
            return errors, None

        return [], {
            "name": name,
            "company_name": get("nazwa firmy") or None,
            "nip": get("nip") or None,
            "phone": get("telefon") or None,
            "email": get("email") or None,
            "street": get("ulica") or None,
            "city": get("miasto") or None,
            "postal_code": get("kod pocztowy") or None,
            "payment_terms": payment_terms,
        }
