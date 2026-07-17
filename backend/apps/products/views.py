import csv
import io
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import DecimalField, QuerySet, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook, load_workbook
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.activity.log import log_activity
from apps.activity.models import ActivityLog
from .filters import ProductFilter
from .models import CustomerProductPrice, Product, ProductStock, StockMovement, Warehouse
from .serializers import (
    CustomerProductPriceSerializer,
    ProductSerializer,
    StockMovementListSerializer,
    StockMovementSerializer,
    StockUpdateSerializer,
    WarehouseSerializer,
    WarehouseStockItemSerializer,
)
from apps.users.permissions import HasCompanyPermission, IsCompanyMember
from apps.users.tenant import filter_queryset_for_current_company


def _stock_owner_user(product, request_user):
    return product.user or request_user


class ProductViewSet(viewsets.ModelViewSet):
    """Full CRUD for products in the user's active company."""
    lookup_field = "uuid"

    serializer_class = ProductSerializer
    required_permission = 'can_manage_products'
    read_permission = None  # any company member may list/read products (needed for orders, WZ, production)
    permission_classes = [IsAuthenticated, IsCompanyMember, HasCompanyPermission]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = ProductFilter
    search_fields = ["name", "description", "sku", "barcode"]
    ordering_fields = [
        "name",
        "unit",
        "price_net",
        "price_gross",
        "vat_rate",
        "min_stock_alert",
        "created_at",
        "updated_at",
        "is_active",
    ]
    ordering = ["-created_at"]

    def get_queryset(self) -> QuerySet:
        qs = Product.objects.all().order_by("-created_at")
        qs = filter_queryset_for_current_company(qs, self.request.user)
        return qs.annotate(
            _stock_total=Coalesce(
                Sum("stocks__quantity_available"),
                Value(Decimal("0")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
        )

    def perform_create(self, serializer):
        serializer.save(
            company=self.request.user.current_company,
            user=self.request.user,
        )

    def perform_update(self, serializer):
        serializer.save(
            company=self.request.user.current_company,
            user=self.request.user,
        )

    @action(detail=True, methods=["post"], url_path="update-stock")
    def update_stock(self, request, uuid=None):
        product = self.get_object()
        input_serializer = StockUpdateSerializer(
            data=request.data,
            context={"product": product, "user": request.user},
        )
        input_serializer.is_valid(raise_exception=True)
        data = input_serializer.validated_data
        existing = getattr(input_serializer, "_existing_movement", None)

        warehouse = get_object_or_404(
            Warehouse.objects.filter(company_id=product.company_id),
            uuid=data["warehouse_id"],
        )
        qty_change: Decimal = data["quantity_change"]
        movement_type = data.get("movement_type", StockMovement.MovementType.ADJUSTMENT)
        ref_type = (data.get("reference_type") or "").strip() or None
        ref_id = data.get("reference_id")
        notes = data.get("notes", "") or ""

        with transaction.atomic():
            stock = (
                ProductStock.objects.select_for_update()
                .filter(product=product, warehouse=warehouse)
                .first()
            )

            if existing:
                if stock is None:
                    return Response(
                        {
                            "detail": (
                                "Cannot update a movement without a matching "
                                "ProductStock row."
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                qty_before = stock.quantity_available - existing.quantity
                new_available = qty_before + qty_change
            else:
                qty_before = stock.quantity_available if stock else Decimal("0")
                new_available = qty_before + qty_change

            if new_available < 0 and not warehouse.allow_negative_stock:
                raise ValidationError(
                    {
                        "quantity_change": (
                            "Resulting quantity_available would be negative "
                            "for this warehouse."
                        )
                    }
                )

            if stock is None:
                stock = ProductStock(
                    company=product.company,
                    product=product,
                    warehouse=warehouse,
                    quantity_available=new_available,
                    quantity_reserved=Decimal("0"),
                )
                stock.save()
            else:
                stock.quantity_available = new_available
                stock.save(update_fields=["quantity_available"])

            movement_user = product.user or request.user

            if existing:
                existing.quantity = qty_change
                existing.quantity_before = qty_before
                existing.quantity_after = new_available
                existing.movement_type = movement_type
                existing.reference_type = ref_type
                existing.reference_id = ref_id
                existing.notes = notes
                existing.created_by = request.user
                existing.save(update_fields=[
                    "quantity",
                    "quantity_before",
                    "quantity_after",
                    "movement_type",
                    "reference_type",
                    "reference_id",
                    "notes",
                    "created_by",
                ])
                movement = existing
                http_status = status.HTTP_200_OK
            else:
                movement = StockMovement.objects.create(
                    company=product.company,
                    product=product,
                    warehouse=warehouse,
                    user=movement_user,
                    movement_type=movement_type,
                    quantity=qty_change,
                    quantity_before=qty_before,
                    quantity_after=new_available,
                    reference_type=ref_type,
                    reference_id=ref_id,
                    notes=notes,
                    created_by=request.user,
                )
                http_status = status.HTTP_201_CREATED

        return Response(
            StockMovementSerializer(movement).data,
            status=http_status,
        )

    # ── Import ────────────────────────────────────────────────────────────────

    _IMPORT_HEADERS = ["Nazwa", "Jednostka", "Cena brutto", "VAT (%)", "SKU", "Kod kreskowy", "Opis", "Alert minimalny"]
    _VALID_VAT_RATES = {Decimal("0"), Decimal("5"), Decimal("8"), Decimal("23")}

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        """GET /api/products/import-template/ — download a blank XLSX template."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Produkty"
        ws.append(self._IMPORT_HEADERS)
        ws.append(["Chleb pszenny", "szt", 2.50, 5, "SKU-001", "", "Chleb na co dzień", ""])

        # Column widths
        for col, width in zip("ABCDEFGH", [30, 12, 14, 10, 15, 16, 30, 16]):
            ws.column_dimensions[col].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="szablon_produkty.xlsx"'
        return resp

    @action(detail=False, methods=["post"], url_path="import", parser_classes=[MultiPartParser])
    def import_products(self, request):
        """POST /api/products/import/ — import products from CSV or XLSX.

        Form fields:
          file     — CSV (semicolon-delimited, UTF-8) or XLSX
          dry_run  — "true" (default) to validate only; "false" to commit

        Dedup logic:
          - SKU present → match existing product by SKU; update if found
          - No SKU → match by name; skip if found (don't overwrite)
        """
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"detail": "Brak pliku."}, status=status.HTTP_400_BAD_REQUEST)

        dry_run = request.data.get("dry_run", "true").lower() != "false"
        filename = file_obj.name.lower()

        try:
            rows = self._parse_import_file(file_obj, filename)
        except Exception as exc:
            return Response({"detail": f"Nie udało się odczytać pliku: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

        errors = []
        valid_products = []

        for i, raw in enumerate(rows, start=2):  # row 1 = header
            row_errors, product_data = self._validate_import_row(raw, i)
            if row_errors:
                errors.extend(row_errors)
            else:
                valid_products.append(product_data)

        if dry_run or errors:
            # Preview: resolve what would happen per row
            company = request.user.current_company
            sku_index = {
                p.sku.lower(): p
                for p in Product.objects.filter(company=company)
                if p.sku
            }
            name_index = {
                p.name.lower(): p
                for p in Product.objects.filter(company=company)
            }
            to_create, to_update = [], []
            for data in valid_products:
                sku = data.get("sku", "").lower()
                name = data.get("name", "").lower()
                if (sku and sku in sku_index) or (name in name_index):
                    to_update.append(data)
                else:
                    to_create.append(data)

            return Response({
                "dry_run": True,
                "to_create": len(to_create),
                "to_update": len(to_update),
                "to_skip": 0,
                "valid_count": len(valid_products),
                "error_count": len(errors),
                "errors": errors,
            })

        # Commit
        company = request.user.current_company
        sku_index = {
            p.sku.lower(): p
            for p in Product.objects.filter(company=company)
            if p.sku
        }
        name_index = {
            p.name.lower(): p
            for p in Product.objects.filter(company=company)
        }

        created = updated = 0
        _UPDATE_FIELDS = ["name", "unit", "price_gross", "price_net", "vat_rate", "sku", "barcode", "description", "min_stock_alert"]

        with transaction.atomic():
            for data in valid_products:
                sku = data.get("sku", "").lower()
                name = data.get("name", "").lower()

                existing = None
                if sku and sku in sku_index:
                    existing = sku_index[sku]
                elif name in name_index:
                    existing = name_index[name]

                if existing:
                    for field in _UPDATE_FIELDS:
                        setattr(existing, field, data[field])
                    existing.save(update_fields=_UPDATE_FIELDS)
                    updated += 1
                else:
                    Product.objects.create(company=company, user=request.user, **data)
                    created += 1

        log_activity(
            user=request.user,
            action="product.import",
            status=ActivityLog.STATUS_SUCCESS,
            object_type="product",
            object_id=f"created={created} updated={updated}",
        )
        return Response({
            "dry_run": False,
            "created": created,
            "updated": updated,
            "skipped": 0,
            "error_count": 0,
            "errors": [],
        }, status=status.HTTP_201_CREATED)

    def _parse_import_file(self, file_obj, filename: str) -> list[dict]:
        """Return a list of raw dicts (header → value) from a CSV or XLSX file."""
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            wb = load_workbook(file_obj, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
            result = []
            for row in rows_iter:
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                result.append({headers[j]: (str(row[j]).strip() if row[j] is not None else "") for j in range(len(headers))})
            return result
        else:
            # CSV — try semicolon first (Polish Excel), then comma
            raw_bytes = file_obj.read()
            text = raw_bytes.decode("utf-8-sig")  # strips BOM
            dialect = "excel" if "," in text.splitlines()[0] and ";" not in text.splitlines()[0] else None
            reader = csv.DictReader(io.StringIO(text), delimiter=";" if dialect is None else ",")
            return [{k.strip(): (v or "").strip() for k, v in row.items() if k is not None} for row in reader]

    def _validate_import_row(self, raw: dict, row_num: int) -> tuple[list[dict], dict | None]:
        """Validate one row. Returns (errors, product_data_or_None)."""
        # Normalise keys — case-insensitive, strip whitespace
        norm = {k.strip().lower(): v.strip() for k, v in raw.items() if k}

        def get(*keys):
            for k in keys:
                if k.lower() in norm:
                    return norm[k.lower()]
            return ""

        errors = []

        name = get("nazwa")
        if not name:
            errors.append({"row": row_num, "field": "Nazwa", "message": "Pole wymagane."})

        unit = get("jednostka")
        if not unit:
            errors.append({"row": row_num, "field": "Jednostka", "message": "Pole wymagane."})

        price_gross_raw = get("cena brutto")
        price_gross = None
        if not price_gross_raw:
            errors.append({"row": row_num, "field": "Cena brutto", "message": "Pole wymagane."})
        else:
            try:
                price_gross = Decimal(price_gross_raw.replace(",", "."))
                if price_gross < 0:
                    raise ValueError
            except (InvalidOperation, ValueError):
                errors.append({"row": row_num, "field": "Cena brutto", "message": f"Nieprawidłowa wartość: '{price_gross_raw}'."})

        vat_raw = get("vat (%)", "vat")
        vat_rate = None
        if not vat_raw:
            errors.append({"row": row_num, "field": "VAT (%)", "message": "Pole wymagane."})
        else:
            try:
                vat_rate = Decimal(vat_raw.replace(",", ".").replace("%", ""))
                if vat_rate not in self._VALID_VAT_RATES:
                    raise ValueError
            except (InvalidOperation, ValueError):
                errors.append({"row": row_num, "field": "VAT (%)", "message": f"Dozwolone wartości: 0, 5, 8, 23. Podano: '{vat_raw}'."})

        if errors:
            return errors, None

        price_net = (price_gross / (1 + vat_rate / 100)).quantize(Decimal("0.01"))

        min_stock_raw = get("alert minimalny")
        min_stock = None
        if min_stock_raw:
            try:
                min_stock = Decimal(min_stock_raw.replace(",", "."))
                if min_stock < 0:
                    raise ValueError
            except (InvalidOperation, ValueError):
                errors.append({"row": row_num, "field": "Alert minimalny", "message": f"Nieprawidłowa wartość: '{min_stock_raw}'."})
                return errors, None

        return [], {
            "name": name,
            "unit": unit,
            "price_gross": price_gross,
            "price_net": price_net,
            "vat_rate": vat_rate,
            "sku": get("sku") or "",
            "barcode": get("kod kreskowy") or "",
            "description": get("opis") or "",
            "min_stock_alert": min_stock if min_stock is not None else Decimal("0"),
        }

    def destroy(self, request, *args, **kwargs):
        """DELETE /api/products/{uuid}/ — only allowed when product has no history."""
        product = self.get_object()
        blockers = self._get_delete_blockers(product)
        if blockers:
            return Response(
                {"detail": "Nie można usunąć produktu z historią.", "blockers": blockers},
                status=status.HTTP_409_CONFLICT,
            )
        product_name = product.name
        product.delete()
        log_activity(
            user=request.user,
            action="product.delete",
            status=ActivityLog.STATUS_SUCCESS,
            object_type="product",
            object_id=product_name,
        )
        return Response(status=status.HTTP_204_NO_CONTENT)

    @staticmethod
    def _get_delete_blockers(product: Product) -> list[str]:
        """Return human-readable reasons why this product cannot be deleted."""
        blockers = []
        if StockMovement.objects.filter(product=product).exists():
            blockers.append("Produkt ma ruchy magazynowe.")
        from apps.orders.models import OrderItem
        if OrderItem.objects.filter(product=product).exists():
            blockers.append("Produkt jest użyty w zamówieniach.")
        from apps.delivery.models import DeliveryItem
        if DeliveryItem.objects.filter(product=product).exists():
            blockers.append("Produkt jest użyty w dokumentach WZ/PZ.")
        from apps.invoices.models import InvoiceItem
        if InvoiceItem.objects.filter(product=product).exists():
            blockers.append("Produkt jest użyty na fakturach.")
        from apps.production.models import RecipeItem
        if RecipeItem.objects.filter(ingredient=product).exists():
            blockers.append("Produkt jest użyty w recepturach.")
        if ProductStock.objects.filter(product=product, quantity_available__gt=0).exists():
            blockers.append("Produkt ma stan magazynowy > 0.")
        return blockers

    @action(detail=False, methods=["get"], url_path="stock-snapshot")
    def stock_snapshot(self, request):
        """Current stock in one warehouse (only lines with ``quantity_available`` > 0)."""
        warehouse_id = request.query_params.get("warehouse_id")
        if not warehouse_id:
            raise ValidationError({"warehouse_id": "This field is required."})
        wh_qs = filter_queryset_for_current_company(
            Warehouse.objects.filter(uuid=warehouse_id),
            request.user,
        )
        warehouse = get_object_or_404(wh_qs)
        items = []
        for ps in (
            ProductStock.objects.filter(
                warehouse=warehouse, company_id=warehouse.company_id
            )
            .select_related("product")
            .order_by("product__name")
        ):
            if ps.quantity_available <= 0:
                continue
            p = ps.product
            items.append(
                {
                    "product_id": str(p.uuid),
                    "product_name": p.name,
                    "sku": p.sku,
                    "unit": p.unit,
                    "quantity_available": f"{ps.quantity_available:.3f}",
                }
            )
        return Response(
            {
                "warehouse_id": str(warehouse.uuid),
                "warehouse_name": warehouse.name,
                "items": items,
            }
        )


class WarehouseViewSet(viewsets.ModelViewSet):
    """Full CRUD for warehouses in the user's active company."""
    lookup_field = "uuid"

    serializer_class = WarehouseSerializer
    required_permission = 'can_manage_warehouses'
    read_permission = None  # any company member may list/read warehouses
    permission_classes = [IsAuthenticated, IsCompanyMember, HasCompanyPermission]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = [
        "code",
        "name",
        "warehouse_type",
        "is_active",
        "allow_negative_stock",
        "fifo_enabled",
    ]
    search_fields = ["code", "name", "address"]
    ordering_fields = [
        "code",
        "name",
        "warehouse_type",
        "created_at",
        "updated_at",
        "is_active",
    ]
    ordering = ["code"]

    def get_queryset(self) -> QuerySet:
        qs = Warehouse.objects.all().order_by("code")
        return filter_queryset_for_current_company(qs, self.request.user)

    def perform_create(self, serializer):
        instance = serializer.save(
            company=self.request.user.current_company,
            user=self.request.user,
        )
        log_activity(
            user=self.request.user,
            action="warehouse.create",
            status=ActivityLog.STATUS_SUCCESS,
            object_type="warehouse",
            object_id=instance.name,
        )

    def perform_update(self, serializer):
        instance = serializer.save(
            company=self.request.user.current_company,
            user=self.request.user,
        )
        log_activity(
            user=self.request.user,
            action="warehouse.update",
            status=ActivityLog.STATUS_SUCCESS,
            object_type="warehouse",
            object_id=instance.name,
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        name = instance.name
        instance.delete()
        log_activity(
            user=request.user,
            action="warehouse.delete",
            status=ActivityLog.STATUS_SUCCESS,
            object_type="warehouse",
            object_id=name,
        )
        return Response(status=status.HTTP_204_NO_CONTENT)

    _STOCK_IMPORT_HEADERS = ["Nazwa produktu", "SKU", "Kod magazynu", "Ilość", "Notatka"]

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        """GET /api/warehouses/import-template/ — download a blank stock import XLSX template."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Stan magazynowy"
        ws.append(self._STOCK_IMPORT_HEADERS)
        ws.append(["Mąka pszenna typ 650", "SKU-001", "MG", 150, "Stan otwarcia"])
        ws.append(["Chleb pszenny", "SKU-002", "MV1", 20, ""])
        ws.append(["Mąka pszenna typ 650", "SKU-001", "MV2", 30, "Stan otwarcia"])

        for col, width in zip("ABCDE", [35, 15, 16, 12, 30]):
            ws.column_dimensions[col].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="szablon_stan_magazynowy.xlsx"'
        return resp

    @action(detail=False, methods=["post"], url_path="import", parser_classes=[MultiPartParser])
    def import_stock(self, request):
        """POST /api/warehouses/import/ — bulk set opening stock from CSV or XLSX.

        Columns: Nazwa produktu | SKU | Kod magazynu | Ilość | Notatka (optional)

        Match product by SKU (preferred) then by name.
        Match warehouse by code.
        Creates/adds a stock adjustment movement for each valid row.
        Dedup: re-importing the same product+warehouse adds to existing stock (not replaces).
        """
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"detail": "Brak pliku."}, status=status.HTTP_400_BAD_REQUEST)

        dry_run = request.data.get("dry_run", "true").lower() != "false"
        filename = file_obj.name.lower()

        try:
            rows = self._parse_stock_import_file(file_obj, filename)
        except Exception as exc:
            return Response({"detail": f"Nie udało się odczytać pliku: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

        company = request.user.current_company

        # Build lookup indexes
        sku_index = {
            p.sku.lower(): p
            for p in Product.objects.filter(company=company)
            if p.sku
        }
        name_index = {
            p.name.lower(): p
            for p in Product.objects.filter(company=company)
        }
        warehouse_index = {
            w.code.lower(): w
            for w in Warehouse.objects.filter(company=company)
        }

        errors = []
        valid_rows = []

        for i, raw in enumerate(rows, start=2):
            row_errors, row_data = self._validate_stock_row(raw, i, sku_index, name_index, warehouse_index)
            if row_errors:
                errors.extend(row_errors)
            else:
                valid_rows.append(row_data)

        if dry_run or errors:
            return Response({
                "dry_run": True,
                "to_create": len(valid_rows),
                "to_update": 0,
                "to_skip": 0,
                "valid_count": len(valid_rows),
                "error_count": len(errors),
                "errors": errors,
            })

        # Commit — create stock movements
        imported = 0
        with transaction.atomic():
            for row in valid_rows:
                product: Product = row["product"]
                warehouse: Warehouse = row["warehouse"]
                qty_change: Decimal = row["quantity"]
                notes: str = row["notes"]

                stock = (
                    ProductStock.objects.select_for_update()
                    .filter(product=product, warehouse=warehouse)
                    .first()
                )
                qty_before = stock.quantity_available if stock else Decimal("0")
                new_available = qty_before + qty_change

                if stock is None:
                    stock = ProductStock(
                        company=company,
                        product=product,
                        warehouse=warehouse,
                        quantity_available=new_available,
                        quantity_reserved=Decimal("0"),
                    )
                    stock.save()
                else:
                    stock.quantity_available = new_available
                    stock.save(update_fields=["quantity_available"])

                StockMovement.objects.create(
                    company=company,
                    product=product,
                    warehouse=warehouse,
                    user=product.user or request.user,
                    movement_type=StockMovement.MovementType.ADJUSTMENT,
                    quantity=qty_change,
                    quantity_before=qty_before,
                    quantity_after=new_available,
                    reference_type="import",
                    notes=notes,
                    created_by=request.user,
                )
                imported += 1

        log_activity(
            user=request.user,
            action="warehouse.import",
            status=ActivityLog.STATUS_SUCCESS,
            object_type="warehouse",
            object_id=f"rows={imported}",
        )
        return Response({
            "dry_run": False,
            "created": imported,
            "updated": 0,
            "skipped": 0,
            "error_count": 0,
            "errors": [],
        }, status=status.HTTP_201_CREATED)

    def _parse_stock_import_file(self, file_obj, filename: str) -> list[dict]:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            wb = load_workbook(file_obj, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
            result = []
            for row in rows_iter:
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                result.append({headers[j]: (str(row[j]).strip() if row[j] is not None else "") for j in range(len(headers))})
            return result
        else:
            raw_bytes = file_obj.read()
            text = raw_bytes.decode("utf-8-sig")
            dialect = "excel" if "," in text.splitlines()[0] and ";" not in text.splitlines()[0] else None
            reader = csv.DictReader(io.StringIO(text), delimiter=";" if dialect is None else ",")
            return [{k.strip(): (v or "").strip() for k, v in row.items() if k is not None} for row in reader]

    def _validate_stock_row(self, raw: dict, row_num: int, sku_index: dict, name_index: dict, warehouse_index: dict):
        norm = {k.strip().lower(): v.strip() for k, v in raw.items() if k}

        def get(*keys):
            for k in keys:
                if k.lower() in norm:
                    return norm[k.lower()]
            return ""

        errors = []

        # Resolve product
        sku = get("sku")
        product_name = get("nazwa produktu", "nazwa")
        product = None
        if sku and sku.lower() in sku_index:
            product = sku_index[sku.lower()]
        elif product_name and product_name.lower() in name_index:
            product = name_index[product_name.lower()]
        else:
            label = sku or product_name or "(brak)"
            errors.append({"row": row_num, "field": "Nazwa produktu / SKU", "message": f"Nie znaleziono produktu: '{label}'."})

        # Resolve warehouse
        wh_code = get("kod magazynu", "kod")
        warehouse = None
        if not wh_code:
            errors.append({"row": row_num, "field": "Kod magazynu", "message": "Pole wymagane."})
        elif wh_code.lower() not in warehouse_index:
            errors.append({"row": row_num, "field": "Kod magazynu", "message": f"Nie znaleziono magazynu o kodzie: '{wh_code}'."})
        else:
            warehouse = warehouse_index[wh_code.lower()]

        # Validate quantity
        qty_raw = get("ilość", "ilosc", "quantity")
        quantity = None
        if not qty_raw:
            errors.append({"row": row_num, "field": "Ilość", "message": "Pole wymagane."})
        else:
            try:
                quantity = Decimal(qty_raw.replace(",", "."))
                if quantity <= 0:
                    raise ValueError
            except (InvalidOperation, ValueError):
                errors.append({"row": row_num, "field": "Ilość", "message": f"Nieprawidłowa wartość: '{qty_raw}'. Musi być liczbą dodatnią."})

        if errors:
            return errors, None

        return [], {
            "product": product,
            "warehouse": warehouse,
            "quantity": quantity,
            "notes": get("notatka") or "Stan otwarcia",
        }

    @action(detail=True, methods=["post"], url_path="transfer")
    def transfer(self, request, uuid=None):
        """POST /api/warehouses/{id}/transfer/ — move stock from this warehouse to another.

        Body:
          destination_warehouse_id  — UUID of target warehouse
          items                     — list of {product_id, quantity}
          notes                     — optional note (applied to all movements)

        Creates two TRANSFER movements per item (deduct source, add destination).
        Validates source has enough stock (unless allow_negative_stock is set).
        """
        source = self.get_object()
        company = request.user.current_company

        dest_id = request.data.get("destination_warehouse_id", "")
        if not dest_id:
            return Response({"detail": "Pole destination_warehouse_id jest wymagane."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            destination = Warehouse.objects.get(uuid=dest_id, company=company)
        except Warehouse.DoesNotExist:
            return Response({"detail": "Nie znaleziono magazynu docelowego."}, status=status.HTTP_400_BAD_REQUEST)

        if source.pk == destination.pk:
            return Response({"detail": "Magazyn źródłowy i docelowy nie mogą być takie same."}, status=status.HTTP_400_BAD_REQUEST)

        items = request.data.get("items", [])
        if not items or not isinstance(items, list):
            return Response({"detail": "Lista produktów jest wymagana."}, status=status.HTTP_400_BAD_REQUEST)

        notes = (request.data.get("notes") or "").strip() or f"Przesunięcie {source.code} → {destination.code}"

        # Validate all items before touching the DB
        validated = []
        for i, item in enumerate(items):
            product_id = item.get("product_id", "")
            qty_raw = item.get("quantity")
            try:
                product = Product.objects.get(uuid=product_id, company=company)
            except Product.DoesNotExist:
                return Response({"detail": f"Produkt nie znaleziony: {product_id}"}, status=status.HTTP_400_BAD_REQUEST)
            try:
                qty = Decimal(str(qty_raw))
                if qty <= 0:
                    raise ValueError
            except (InvalidOperation, ValueError, TypeError):
                return Response({"detail": f"Nieprawidłowa ilość dla produktu {product.name}."}, status=status.HTTP_400_BAD_REQUEST)

            source_stock = ProductStock.objects.filter(product=product, warehouse=source).first()
            available = source_stock.quantity_available if source_stock else Decimal("0")
            if qty > available and not source.allow_negative_stock:
                return Response(
                    {"detail": f"Niewystarczający stan dla '{product.name}': dostępne {available}, żądane {qty}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            validated.append({"product": product, "quantity": qty, "source_stock": source_stock})

        transferred = 0
        with transaction.atomic():
            for v in validated:
                product = v["product"]
                qty = v["quantity"]

                # --- Deduct from source ---
                src_stock = (
                    ProductStock.objects.select_for_update()
                    .filter(product=product, warehouse=source)
                    .first()
                )
                src_before = src_stock.quantity_available if src_stock else Decimal("0")
                src_after = src_before - qty
                if src_stock:
                    src_stock.quantity_available = src_after
                    src_stock.save(update_fields=["quantity_available"])
                else:
                    src_stock = ProductStock.objects.create(
                        company=company, product=product, warehouse=source,
                        quantity_available=src_after, quantity_reserved=Decimal("0"),
                    )
                StockMovement.objects.create(
                    company=company, product=product, warehouse=source,
                    user=product.user or request.user,
                    movement_type=StockMovement.MovementType.TRANSFER,
                    quantity=-qty,
                    quantity_before=src_before, quantity_after=src_after,
                    reference_type="transfer",
                    notes=notes, created_by=request.user,
                )

                # --- Add to destination ---
                dst_stock = (
                    ProductStock.objects.select_for_update()
                    .filter(product=product, warehouse=destination)
                    .first()
                )
                dst_before = dst_stock.quantity_available if dst_stock else Decimal("0")
                dst_after = dst_before + qty
                if dst_stock:
                    dst_stock.quantity_available = dst_after
                    dst_stock.save(update_fields=["quantity_available"])
                else:
                    ProductStock.objects.create(
                        company=company, product=product, warehouse=destination,
                        quantity_available=dst_after, quantity_reserved=Decimal("0"),
                    )
                StockMovement.objects.create(
                    company=company, product=product, warehouse=destination,
                    user=product.user or request.user,
                    movement_type=StockMovement.MovementType.TRANSFER,
                    quantity=qty,
                    quantity_before=dst_before, quantity_after=dst_after,
                    reference_type="transfer",
                    notes=notes, created_by=request.user,
                )
                transferred += 1

        log_activity(
            user=request.user,
            action="warehouse.transfer",
            status=ActivityLog.STATUS_SUCCESS,
            object_type="warehouse",
            object_id=f"{source.code}→{destination.code} products={transferred}",
        )
        return Response({"transferred": transferred, "source": source.code, "destination": destination.code})

    @action(detail=True, methods=["get"], url_path="stock")
    def stock(self, request, uuid=None):
        """GET /api/warehouses/{id}/stock/ — all ProductStock rows for this warehouse.

        Query params:
          ?below_minimum=true  — only items below min_stock_alert
          ?search=name         — filter by product name (icontains)
        """
        warehouse = self.get_object()
        company = request.user.current_company

        qs = (
            ProductStock.objects.filter(
                warehouse=warehouse,
                company_id=company.id,
                product__is_active=True,
            )
            .select_related("product")
            .order_by("product__name")
        )

        search = request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(product__name__icontains=search)

        below_minimum = request.query_params.get("below_minimum", "").lower() == "true"
        if below_minimum:
            qs = [
                s for s in qs
                if s.product.min_stock_alert and s.quantity_total < s.product.min_stock_alert
            ]

        serializer = WarehouseStockItemSerializer(qs, many=True)
        return Response(serializer.data)


class StockMovementViewSet(viewsets.ReadOnlyModelViewSet):
    """Read-only list/detail of StockMovements for the current company.

    Filter params: ?product={uuid}, ?warehouse={uuid}, ?type={TYPE},
                   ?date_from=YYYY-MM-DD, ?date_to=YYYY-MM-DD
    """
    lookup_field = "uuid"

    serializer_class = StockMovementListSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ["created_at"]
    ordering = ["-created_at"]

    def get_queryset(self) -> QuerySet:
        company = self.request.user.current_company
        qs = (
            StockMovement.objects.filter(company=company)
            .select_related("product", "warehouse", "created_by")
        )

        product_id = self.request.query_params.get("product")
        warehouse_id = self.request.query_params.get("warehouse")
        movement_type = self.request.query_params.get("type")
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")

        if product_id:
            qs = qs.filter(product__uuid=product_id)
        if warehouse_id:
            qs = qs.filter(warehouse__uuid=warehouse_id)
        if movement_type:
            qs = qs.filter(movement_type=movement_type)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        return qs


class CustomerProductPriceViewSet(viewsets.ModelViewSet):
    serializer_class = CustomerProductPriceSerializer
    required_permission = 'can_manage_customers'
    permission_classes = [IsAuthenticated, IsCompanyMember, HasCompanyPermission]
    pagination_class = None
    http_method_names = ["get", "post", "patch", "delete", "head", "options"]

    def get_queryset(self) -> QuerySet:
        qs = CustomerProductPrice.objects.select_related("product", "customer").order_by(
            "product__name"
        )
        qs = filter_queryset_for_current_company(qs, self.request.user)
        customer_id = self.request.query_params.get("customer")
        if customer_id:
            qs = qs.filter(customer_id=customer_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.current_company)

    def perform_update(self, serializer):
        serializer.save(company=self.request.user.current_company)
