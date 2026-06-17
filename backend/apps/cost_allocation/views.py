import csv
import io
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.db import IntegrityError
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.ksef.models import ReceivedKSeFInvoice, ReceivedKSeFInvoiceLine
from apps.users.permissions import IsCompanyMember, ModuleRequired

from .models import CostProject, InvoiceAnnotation, InvoiceLineAnnotation, InvoiceLineAnnotationSplit
from .serializers import CostProjectSerializer, InvoiceAnnotationSerializer

EXPORT_HEADERS = [
    "Nr KSeF",
    "Data faktury",
    "Nr faktury",
    "Dostawca",
    "NIP dostawcy",
    "Waluta",
    "Pozycja",
    "Wartość netto",
    "Kwota VAT",
    "Wartość brutto",
    "Stawka VAT %",
    "Projekt",
    "Udział %",
    "Ilość (podział)",
    "Prywatne",
    "Uwagi do pozycji",
    "Notatki księgowe",
    "Status",
]

STATUS_LABELS = {
    "pending": "Do opisania",
    "annotated": "Opisana",
    "exported": "Wyeksportowana",
    "booked": "Zaksięgowana",
}


def _calc_vat(line_net, vat_rate_str):
    """Return (vat_amount, gross) as Decimal, or (None, None) if rate is non-numeric (e.g. 'zw')."""
    try:
        rate = Decimal(str(vat_rate_str).replace("%", "").strip())
        vat = (line_net * rate / 100).quantize(Decimal("0.01"))
        return vat, line_net + vat
    except (InvalidOperation, TypeError):
        return None, None


def _build_rows(invoices_qs):
    """Yield (inv_id, should_mark_exported, row_list) for every split/line."""
    for inv in invoices_qs:
        try:
            inv_ann = inv.annotation
            notes = inv_ann.accounting_notes
            acc_status = STATUS_LABELS.get(inv_ann.accounting_status, inv_ann.accounting_status)
            mark_exported = inv_ann.accounting_status != InvoiceAnnotation.STATUS_BOOKED
        except InvoiceAnnotation.DoesNotExist:
            notes = ""
            acc_status = STATUS_LABELS["pending"]
            mark_exported = True

        lines = list(inv.lines.prefetch_related("annotation__splits__project", "annotation__project").all())
        if not lines:
            yield inv.id, mark_exported, [
                inv.ksef_number or "",
                str(inv.issue_date) if inv.issue_date else "",
                inv.invoice_number or "",
                inv.seller_name or "",
                inv.seller_nip or "",
                inv.currency or "",
                "", "", "", "", "", "", "", "", "", notes, acc_status,
            ]
            continue

        for line in lines:
            vat_amt, gross = _calc_vat(line.line_net, line.vat_rate)
            try:
                la = line.annotation
                is_private = "TAK" if la.is_private else ""
                line_note = la.note
                splits = list(la.splits.all())
                # Back-compat: no splits but legacy project
                if not splits and la.project_id:
                    splits_data = [(la.project.name if la.project else "", Decimal("100"), None, "")]
                elif splits:
                    splits_data = [
                        (s.project.name if s.project else "", s.percentage, s.quantity, s.note)
                        for s in splits
                    ]
                else:
                    splits_data = [("", Decimal("100"), None, "")]
            except InvoiceLineAnnotation.DoesNotExist:
                is_private = ""
                line_note = ""
                splits_data = [("", Decimal("100"), None, "")]

            for project_name, pct, split_qty, split_note in splits_data:
                factor = pct / Decimal("100")
                split_net = (line.line_net * factor).quantize(Decimal("0.01"))
                split_vat = (vat_amt * factor).quantize(Decimal("0.01")) if vat_amt is not None else ""
                split_gross = (gross * factor).quantize(Decimal("0.01")) if gross is not None else ""
                pct_str = f"{pct:.0f}%" if pct != Decimal("100") else ""

                yield inv.id, mark_exported, [
                    inv.ksef_number or "",
                    str(inv.issue_date) if inv.issue_date else "",
                    inv.invoice_number or "",
                    inv.seller_name or "",
                    inv.seller_nip or "",
                    inv.currency or "",
                    line.name or "",
                    split_net,
                    split_vat,
                    split_gross,
                    line.vat_rate or "",
                    project_name,
                    pct_str,
                    split_qty if split_qty is not None else "",
                    is_private,
                    split_note or line_note,
                    notes,
                    acc_status,
                ]


def _mark_exported(ids):
    if not ids:
        return
    now = timezone.now()
    for inv_id in ids:
        InvoiceAnnotation.objects.update_or_create(
            invoice_id=inv_id,
            defaults={
                "accounting_status": InvoiceAnnotation.STATUS_EXPORTED,
                "exported_at": now,
            },
        )


class CostProjectListView(APIView):
    """
    GET  /api/cost-allocation/projects/  — list active projects for current company
    POST /api/cost-allocation/projects/  — create a new project
    """

    permission_classes = [IsAuthenticated, IsCompanyMember, ModuleRequired]
    module_required = "cost_allocation"

    def get(self, request):
        company = request.user.current_company
        projects = CostProject.objects.filter(company=company, is_active=True).order_by("name")
        return Response(CostProjectSerializer(projects, many=True).data)

    def post(self, request):
        company = request.user.current_company
        serializer = CostProjectSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        try:
            serializer.save(company=company)
        except IntegrityError:
            return Response(
                {"name": ["Projekt o tej nazwie już istnieje."]},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class CostProjectDetailView(APIView):
    """
    PATCH  /api/cost-allocation/projects/<id>/  — update project
    DELETE /api/cost-allocation/projects/<id>/  — soft-delete (is_active=False)
    """

    permission_classes = [IsAuthenticated, IsCompanyMember, ModuleRequired]
    module_required = "cost_allocation"

    def _get_project(self, company, pk):
        return CostProject.objects.filter(id=pk, company=company).first()

    def patch(self, request, pk):
        company = request.user.current_company
        project = self._get_project(company, pk)
        if not project:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        serializer = CostProjectSerializer(project, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        try:
            serializer.save()
        except IntegrityError:
            return Response(
                {"name": ["Projekt o tej nazwie już istnieje."]},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response(serializer.data)

    def delete(self, request, pk):
        company = request.user.current_company
        project = self._get_project(company, pk)
        if not project:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        project.is_active = False
        project.save(update_fields=["is_active"])
        return Response(status=status.HTTP_204_NO_CONTENT)


class InvoiceAnnotationView(APIView):
    """
    GET   /api/cost-allocation/invoices/<ksef_number>/annotation/
          Returns current annotation + per-line annotations keyed by position.
          Returns {} (empty object) if no annotation exists yet — not 404.

    PATCH /api/cost-allocation/invoices/<ksef_number>/annotation/
          Upserts invoice annotation and optionally all line annotations in one request.
          Body: {
            accountingStatus?, accountingNotes?,
            lineAnnotations?: { "<position>": { project?, isPrivate?, note? } }
          }
    """

    permission_classes = [IsAuthenticated, IsCompanyMember, ModuleRequired]
    module_required = "cost_allocation"

    def _get_invoice(self, company, ksef_number):
        return ReceivedKSeFInvoice.objects.filter(
            company=company, ksef_number=ksef_number
        ).first()

    def get(self, request, ksef_number):
        company = request.user.current_company
        invoice = self._get_invoice(company, ksef_number)
        if not invoice:
            return Response({"detail": "Faktura nie znaleziona."}, status=status.HTTP_404_NOT_FOUND)
        try:
            ann = invoice.annotation
            return Response(InvoiceAnnotationSerializer(ann).data)
        except InvoiceAnnotation.DoesNotExist:
            return Response({})

    def patch(self, request, ksef_number):
        company = request.user.current_company
        invoice = self._get_invoice(company, ksef_number)
        if not invoice:
            return Response({"detail": "Faktura nie znaleziona."}, status=status.HTTP_404_NOT_FOUND)

        # Upsert invoice-level annotation
        ann, _ = InvoiceAnnotation.objects.get_or_create(invoice=invoice)
        invoice_data = {}
        if "accountingStatus" in request.data:
            invoice_data["accounting_status"] = request.data["accountingStatus"]
        if "accountingNotes" in request.data:
            invoice_data["accounting_notes"] = request.data["accountingNotes"]
        if invoice_data:
            for field, val in invoice_data.items():
                setattr(ann, field, val)
            ann.save()

        # Upsert per-line annotations keyed by position
        line_annotations = request.data.get("lineAnnotations") or {}
        if line_annotations:
            lines_by_position = {str(ln.position): ln for ln in invoice.lines.all()}
            valid_project_ids = set(
                CostProject.objects.filter(company=company, is_active=True).values_list("id", flat=True)
            )

            for pos_str, la_data in line_annotations.items():
                line = lines_by_position.get(str(pos_str))
                if not line:
                    continue

                la, _ = InvoiceLineAnnotation.objects.get_or_create(line=line)
                if "isPrivate" in la_data:
                    la.is_private = bool(la_data["isPrivate"])
                if "note" in la_data:
                    la.note = la_data.get("note", "")
                la.save()

                # Handle splits — full replace when provided
                if "splits" in la_data:
                    la.splits.all().delete()
                    for split in la_data["splits"]:
                        project_id = split.get("project")
                        if project_id and str(project_id) not in {str(p) for p in valid_project_ids}:
                            continue
                        try:
                            pct = Decimal(str(split.get("percentage", 100)))
                        except (InvalidOperation, TypeError):
                            pct = Decimal("100")
                        qty_raw = split.get("quantity")
                        try:
                            qty = Decimal(str(qty_raw)) if qty_raw is not None else None
                        except (InvalidOperation, TypeError):
                            qty = None
                        InvoiceLineAnnotationSplit.objects.create(
                            line_annotation=la,
                            project_id=project_id or None,
                            percentage=pct,
                            quantity=qty,
                            note=split.get("note", ""),
                        )

        ann.refresh_from_db()
        return Response(InvoiceAnnotationSerializer(ann).data)


class CostAllocationExportView(APIView):
    """
    GET /api/cost-allocation/export/?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD&fmt=csv|xlsx
    Export annotated invoice lines. Default format: xlsx.
    Marks exported invoices as 'exported' unless already 'booked'.
    """

    permission_classes = [IsAuthenticated, IsCompanyMember, ModuleRequired]
    module_required = "cost_allocation"

    def get(self, request):
        company = request.user.current_company
        date_from = request.query_params.get("date_from", "")
        date_to = request.query_params.get("date_to", "")
        fmt = request.query_params.get("fmt", "xlsx").lower()

        invoices_qs = (
            ReceivedKSeFInvoice.objects.filter(company=company)
            .prefetch_related("lines", "lines__annotation__project", "annotation")
            .order_by("issue_date", "invoice_number")
        )
        if date_from:
            invoices_qs = invoices_qs.filter(issue_date__gte=date_from[:10])
        if date_to:
            invoices_qs = invoices_qs.filter(issue_date__lte=date_to[:10])

        rows = list(_build_rows(invoices_qs))
        ids_to_export = [inv_id for inv_id, mark, _ in rows if mark]

        if fmt == "csv":
            response = self._csv_response(rows)
        else:
            response = self._xlsx_response(rows)

        _mark_exported(ids_to_export)
        return response

    def _csv_response(self, rows):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="adnotacje_kosztowe.csv"'
        response.write("\ufeff")      # UTF-8 BOM
        response.write("sep=;\r\n")   # force semicolon in Excel/LibreOffice

        writer = csv.writer(response, delimiter=";", quoting=csv.QUOTE_ALL)
        writer.writerow(EXPORT_HEADERS)
        for _, _, row in rows:
            # Format decimals with comma for Polish locale
            writer.writerow([
                str(v).replace(".", ",") if isinstance(v, Decimal) else v
                for v in row
            ])
        return response

    def _xlsx_response(self, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "Adnotacje kosztowe"

        # Header row styling
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        ws.append(EXPORT_HEADERS)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Numeric column indices (1-based): Wartość netto=8, Kwota VAT=9, Wartość brutto=10
        numeric_cols = {8, 9, 10}

        for _, _, row in rows:
            ws.append(row)
            # Apply number format to amount cells
            excel_row = ws.max_row
            for col_idx in numeric_cols:
                cell = ws.cell(row=excel_row, column=col_idx)
                if isinstance(cell.value, Decimal):
                    cell.value = float(cell.value)
                    cell.number_format = '#,##0.00'

        # Auto-size columns (cap at 50)
        for col_idx, _ in enumerate(EXPORT_HEADERS, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
                default=8,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="adnotacje_kosztowe.xlsx"'
        return response
